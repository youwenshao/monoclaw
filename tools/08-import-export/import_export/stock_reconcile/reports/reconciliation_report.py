"""Generate full reconciliation reports and export to Excel."""

from __future__ import annotations

from typing import Any

from openclaw_shared.database import get_db


class ReconciliationReport:
    """Produces structured reconciliation reports with export capability."""

    @staticmethod
    def generate(db_path: str, shipment_id: int) -> dict[str, Any]:
        with get_db(db_path) as conn:
            shipment = conn.execute(
                "SELECT * FROM shipments WHERE id = ?", (shipment_id,)
            ).fetchone()
            if not shipment:
                raise ValueError(f"Shipment {shipment_id} not found")

            results = conn.execute(
                """SELECT rr.*,
                          mi.item_description AS manifest_description,
                          mi.sku AS manifest_sku, mi.quantity AS manifest_qty_raw,
                          mi.weight_kg, mi.carton_count,
                          ri.item_description AS receipt_description,
                          ri.sku AS receipt_sku,
                          ri.quantity_received AS receipt_qty_raw,
                          ri.condition, ri.damage_notes
                   FROM reconciliation_results rr
                   LEFT JOIN manifest_items mi ON rr.manifest_item_id = mi.id
                   LEFT JOIN receipt_items ri ON rr.receipt_item_id = ri.id
                   WHERE rr.shipment_id = ?
                   ORDER BY rr.id""",
                (shipment_id,),
            ).fetchall()

        shipment_dict = dict(shipment)
        result_list = [dict(r) for r in results]

        summary = {
            "total_items": len(result_list),
            "matched": 0, "shortage": 0, "overage": 0,
            "damaged": 0, "unmatched_manifest": 0, "unmatched_receipt": 0,
            "total_manifest_qty": 0.0,
            "total_received_qty": 0.0,
            "total_variance": 0.0,
        }
        for r in result_list:
            status = r.get("status", "matched")
            summary[status] = summary.get(status, 0) + 1
            summary["total_manifest_qty"] += r.get("quantity_manifest") or 0
            summary["total_received_qty"] += r.get("quantity_received") or 0
            summary["total_variance"] += r.get("variance") or 0

        if summary["total_items"] > 0:
            summary["accuracy_pct"] = round(
                summary["matched"] / summary["total_items"] * 100, 1
            )
        else:
            summary["accuracy_pct"] = 0.0

        return {
            "shipment": shipment_dict,
            "summary": summary,
            "line_items": result_list,
        }

    @staticmethod
    def to_excel(report: dict[str, Any], output_path: str) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        # -- Summary sheet --
        ws_summary = wb.active
        ws_summary.title = "Summary"
        header_font = Font(bold=True, size=12)
        gold_fill = PatternFill(start_color="D4A843", end_color="D4A843", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        shipment = report["shipment"]
        summary = report["summary"]

        ws_summary.append(["Reconciliation Report"])
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.append([])

        info_fields = [
            ("Shipment Reference", shipment.get("shipment_reference")),
            ("B/L Number", shipment.get("bl_number")),
            ("Vessel", shipment.get("vessel_name")),
            ("Voyage", shipment.get("voyage")),
            ("Origin Port", shipment.get("origin_port")),
            ("Arrival Date", shipment.get("arrival_date")),
            ("Load Type", shipment.get("load_type")),
            ("Status", shipment.get("status")),
        ]
        for label, value in info_fields:
            ws_summary.append([label, value])

        ws_summary.append([])
        ws_summary.append(["Reconciliation Summary"])

        stat_fields = [
            ("Total Items", summary["total_items"]),
            ("Matched", summary["matched"]),
            ("Shortages", summary["shortage"]),
            ("Overages", summary["overage"]),
            ("Damaged", summary["damaged"]),
            ("Accuracy %", f"{summary['accuracy_pct']}%"),
            ("Total Manifest Qty", summary["total_manifest_qty"]),
            ("Total Received Qty", summary["total_received_qty"]),
            ("Total Variance", summary["total_variance"]),
        ]
        for label, value in stat_fields:
            ws_summary.append([label, value])

        for col in ["A", "B"]:
            ws_summary.column_dimensions[col].width = 25

        # -- Detail sheet --
        ws_detail = wb.create_sheet("Line Items")
        headers = [
            "Manifest Description", "Manifest SKU", "Manifest Qty",
            "Receipt Description", "Receipt SKU", "Received Qty",
            "Variance", "Status", "Confidence", "Condition", "Damage Notes",
        ]

        for col_idx, h in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = gold_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        status_fills = {
            "matched": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "shortage": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "overage": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "damaged": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }

        for row_idx, item in enumerate(report["line_items"], 2):
            values = [
                item.get("manifest_description"),
                item.get("manifest_sku"),
                item.get("quantity_manifest"),
                item.get("receipt_description"),
                item.get("receipt_sku"),
                item.get("quantity_received"),
                item.get("variance"),
                item.get("status"),
                item.get("match_confidence"),
                item.get("condition"),
                item.get("damage_notes"),
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=v)
                cell.border = thin_border

            status = item.get("status", "")
            if status in status_fills:
                status_cell = ws_detail.cell(row=row_idx, column=8)
                status_cell.fill = status_fills[status]

        for i, _ in enumerate(headers, 1):
            ws_detail.column_dimensions[ws_detail.cell(row=1, column=i).column_letter].width = 18

        wb.save(output_path)
        return output_path

    @staticmethod
    def get_summary_stats(db_path: str) -> dict[str, Any]:
        with get_db(db_path) as conn:
            total_shipments = conn.execute("SELECT COUNT(*) FROM shipments").fetchone()[0]
            reconciled = conn.execute(
                "SELECT COUNT(*) FROM shipments WHERE status IN ('reconciled', 'closed')"
            ).fetchone()[0]
            active_discrepancies = conn.execute(
                "SELECT COUNT(*) FROM reconciliation_results WHERE status != 'matched'"
            ).fetchone()[0]
            pending_receipts = conn.execute(
                """SELECT COUNT(*) FROM shipments s
                   WHERE s.status NOT IN ('reconciled', 'closed')
                   AND NOT EXISTS (
                       SELECT 1 FROM warehouse_receipts wr WHERE wr.shipment_id = s.id
                   )"""
            ).fetchone()[0]

            accuracy_row = conn.execute(
                """SELECT
                       CASE WHEN COUNT(*) > 0
                            THEN ROUND(SUM(CASE WHEN status = 'matched' THEN 1 ELSE 0 END) * 100.0 / COUNT(*), 1)
                            ELSE 0 END AS accuracy
                   FROM reconciliation_results"""
            ).fetchone()

        return {
            "total_shipments": total_shipments,
            "reconciled": reconciled,
            "active_discrepancies": active_discrepancies,
            "pending_receipts": pending_receipts,
            "overall_accuracy_pct": accuracy_row["accuracy"] if accuracy_row else 0,
        }
