"""Parse packing list documents (PDF / Excel) into structured item records."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any


class PackingListParser:
    """Extracts packing list line items from tabular PDF or Excel formats."""

    COLUMN_ALIASES: dict[str, list[str]] = {
        "carton_no": ["carton no", "ctn no", "carton", "box no", "case no", "package no"],
        "description": ["description", "desc", "goods", "commodity", "item", "product"],
        "sku": ["sku", "item code", "part no", "article no", "product code"],
        "quantity": ["quantity", "qty", "pcs", "units"],
        "unit": ["unit", "uom"],
        "net_weight": ["net weight", "n/w", "nw", "net wt"],
        "gross_weight": ["gross weight", "g/w", "gw", "gross wt"],
        "dimensions": ["dimensions", "dims", "l x w x h", "measurement", "cbm"],
    }

    def parse(self, file_path: str) -> list[dict[str, Any]]:
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            return self._parse_pdf(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self._parse_excel(file_path)
        elif suffix == ".csv":
            return self._parse_csv(file_path)
        else:
            raise ValueError(f"Unsupported packing list format: {suffix}")

    def _parse_pdf(self, file_path: str) -> list[dict[str, Any]]:
        import pdfplumber

        all_items: list[dict[str, Any]] = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    text = page.extract_text()
                    if text:
                        all_items.extend(self._parse_text_fallback(text))
                    continue

                for table in tables:
                    if len(table) < 2:
                        continue
                    header_row = [str(c).strip().lower() if c else "" for c in table[0]]
                    col_map = self._map_columns(header_row)
                    if not col_map:
                        continue
                    for row in table[1:]:
                        cells = [str(c).strip() if c else "" for c in row]
                        item = self._extract_item(cells, col_map)
                        if item:
                            all_items.append(item)

        return all_items

    def _parse_excel(self, file_path: str) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
        wb.close()

        return self._parse_rows(rows)

    def _parse_csv(self, file_path: str) -> list[dict[str, Any]]:
        import csv

        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = [[cell.strip() for cell in row] for row in reader]

        return self._parse_rows(rows)

    def _parse_rows(self, rows: list[list[str]]) -> list[dict[str, Any]]:
        header_idx = self._find_header_row(rows)
        if header_idx is None:
            return []

        col_map = self._map_columns([c.lower() for c in rows[header_idx]])
        items: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if not any(cell for cell in row):
                continue
            item = self._extract_item(row, col_map)
            if item:
                items.append(item)
        return items

    def _find_header_row(self, rows: list[list[str]]) -> int | None:
        keywords = {"description", "desc", "sku", "quantity", "qty", "carton", "weight"}
        for idx, row in enumerate(rows):
            cells_lower = {c.lower().strip() for c in row if c}
            if len(cells_lower & keywords) >= 2:
                return idx
        return None

    def _map_columns(self, header_row: list[str]) -> dict[str, int]:
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            cell_clean = cell.strip()
            for field_name, aliases in self.COLUMN_ALIASES.items():
                if cell_clean in aliases or any(a in cell_clean for a in aliases):
                    col_map[field_name] = idx
                    break
        return col_map

    def _extract_item(self, row: list[str], col_map: dict[str, int]) -> dict[str, Any] | None:
        def _get(field: str) -> str:
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return ""

        description = _get("description")
        sku = _get("sku")
        if not description and not sku:
            return None

        return {
            "carton_no": _get("carton_no") or None,
            "description": description,
            "sku": sku or None,
            "quantity": self._to_float(_get("quantity")),
            "unit": (_get("unit") or "pcs").lower(),
            "net_weight_kg": self._to_float(_get("net_weight")),
            "gross_weight_kg": self._to_float(_get("gross_weight")),
            "dimensions": _get("dimensions") or None,
        }

    def _parse_text_fallback(self, text: str) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        qty_pattern = re.compile(r"(\d[\d,.]*)\s*(pcs|units|ctns|sets?|kgs?)", re.I)
        for line in text.splitlines():
            m = qty_pattern.search(line)
            if m:
                items.append({
                    "carton_no": None,
                    "description": line[:m.start()].strip().rstrip("-:"),
                    "sku": None,
                    "quantity": self._to_float(m.group(1)),
                    "unit": m.group(2).lower(),
                    "net_weight_kg": None,
                    "gross_weight_kg": None,
                    "dimensions": None,
                })
        return items

    @staticmethod
    def _to_float(raw: Any) -> float | None:
        if not raw:
            return None
        try:
            return float(str(raw).replace(",", "").strip())
        except (ValueError, TypeError):
            return None
