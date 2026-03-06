"""Parse warehouse receipt documents into structured receipt data."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any


class WarehouseReceiptParser:
    """Extracts warehouse receipt header + received items from PDF, Excel, or CSV."""

    HEADER_PATTERNS = {
        "receipt_number": re.compile(
            r"(?:receipt|grn|goods\s+received)\s*(?:no\.?|number|#)?\s*[:\-]?\s*([A-Z0-9\-/]+)", re.I
        ),
        "received_date": re.compile(
            r"(?:received?\s+date|date\s+received|grn\s+date)\s*[:\-]?\s*(\d{4}[\-/]\d{2}[\-/]\d{2}|\d{2}[\-/]\d{2}[\-/]\d{4})", re.I
        ),
        "warehouse": re.compile(
            r"warehouse\s*(?:name|location)?\s*[:\-]?\s*(.+)", re.I
        ),
        "received_by": re.compile(
            r"received\s+by\s*[:\-]?\s*(.+)", re.I
        ),
    }

    ITEM_ALIASES: dict[str, list[str]] = {
        "description": ["description", "desc", "goods", "item description", "commodity", "item"],
        "sku": ["sku", "item code", "product code", "part no", "article no"],
        "quantity_received": ["quantity", "qty", "qty received", "quantity received", "received qty", "pcs", "units"],
        "unit": ["unit", "uom", "unit of measure"],
        "condition": ["condition", "status", "quality", "remarks"],
    }

    def parse(self, file_path: str) -> dict[str, Any]:
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            return self._parse_pdf(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self._parse_excel(file_path)
        elif suffix == ".csv":
            return self._parse_csv(file_path)
        else:
            raise ValueError(f"Unsupported receipt format: {suffix}")

    def _parse_pdf(self, file_path: str) -> dict[str, Any]:
        import pdfplumber

        text_lines: list[str] = []
        tables: list[list[list[str | None]]] = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_lines.extend(page_text.splitlines())
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)

        header = self._extract_header(text_lines)
        items = self._extract_items_from_tables(tables) if tables else []

        return {**header, "items": items, "source_format": "pdf"}

    def _parse_excel(self, file_path: str) -> dict[str, Any]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
        wb.close()

        header = self._extract_header_from_rows(rows)
        items = self._extract_items_from_rows(rows)

        return {**header, "items": items, "source_format": "excel"}

    def _parse_csv(self, file_path: str) -> dict[str, Any]:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = [[cell.strip() for cell in row] for row in reader]

        header = self._extract_header_from_rows(rows)
        items = self._extract_items_from_rows(rows)

        return {**header, "items": items, "source_format": "csv"}

    # ── Header extraction ──────────────────────────────────────────────────

    def _extract_header(self, lines: list[str]) -> dict[str, str | None]:
        header: dict[str, str | None] = {
            "receipt_number": None,
            "received_date": None,
            "warehouse": None,
            "received_by": None,
        }
        for line in lines[:30]:
            for field, pattern in self.HEADER_PATTERNS.items():
                if not header[field]:
                    m = pattern.search(line)
                    if m:
                        header[field] = m.group(1).strip()
        return header

    def _extract_header_from_rows(self, rows: list[list[str]]) -> dict[str, str | None]:
        flat_lines = [" ".join(row) for row in rows[:20]]
        return self._extract_header(flat_lines)

    # ── Item extraction ────────────────────────────────────────────────────

    def _extract_items_from_tables(self, tables: list[list[list[str | None]]]) -> list[dict[str, Any]]:
        all_items: list[dict[str, Any]] = []
        for table in tables:
            if not table or len(table) < 2:
                continue
            header_row = [str(c).strip().lower() if c else "" for c in table[0]]
            col_map = self._map_columns(header_row)
            if not col_map:
                continue
            for row in table[1:]:
                cells = [str(c).strip() if c else "" for c in row]
                item = self._row_to_item(cells, col_map)
                if item:
                    all_items.append(item)
        return all_items

    def _extract_items_from_rows(self, rows: list[list[str]]) -> list[dict[str, Any]]:
        header_idx = self._find_header_row(rows)
        if header_idx is None:
            return []

        col_map = self._map_columns([c.lower() for c in rows[header_idx]])
        items: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if not any(cell for cell in row):
                continue
            item = self._row_to_item(row, col_map)
            if item:
                items.append(item)
        return items

    def _find_header_row(self, rows: list[list[str]]) -> int | None:
        keywords = {"description", "desc", "sku", "quantity", "qty", "condition"}
        for idx, row in enumerate(rows):
            cells_lower = {c.lower().strip() for c in row if c}
            if len(cells_lower & keywords) >= 2:
                return idx
        return None

    def _map_columns(self, header_row: list[str]) -> dict[str, int]:
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            cell_clean = cell.strip()
            for field_name, aliases in self.ITEM_ALIASES.items():
                if cell_clean in aliases or any(a in cell_clean for a in aliases):
                    col_map[field_name] = idx
                    break
        return col_map

    def _row_to_item(self, row: list[str], col_map: dict[str, int]) -> dict[str, Any] | None:
        def _get(field: str) -> str:
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return ""

        description = _get("description")
        sku = _get("sku")
        if not description and not sku:
            return None

        condition_raw = _get("condition").lower().strip()
        condition = self._normalize_condition(condition_raw)

        return {
            "description": description,
            "sku": sku or None,
            "quantity_received": self._to_float(_get("quantity_received")),
            "unit": (_get("unit") or "pcs").lower(),
            "condition": condition,
        }

    @staticmethod
    def _normalize_condition(raw: str) -> str:
        if not raw or raw in ("good", "ok", "fine", "normal", ""):
            return "good"
        if any(kw in raw for kw in ("damage", "broken", "dent", "wet", "crush")):
            return "damaged"
        if any(kw in raw for kw in ("partial", "short", "missing")):
            return "partial"
        return "good"

    @staticmethod
    def _to_float(raw: Any) -> float | None:
        if not raw:
            return None
        try:
            return float(str(raw).replace(",", "").strip())
        except (ValueError, TypeError):
            return None
