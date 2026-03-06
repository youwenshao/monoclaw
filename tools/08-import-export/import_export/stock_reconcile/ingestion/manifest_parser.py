"""Parse shipping manifests from PDF, Excel, or CSV into structured data."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any


class ManifestParser:
    """Extracts manifest data from PDF, Excel, or CSV shipping documents."""

    FIELD_ALIASES: dict[str, list[str]] = {
        "bl_number": ["b/l", "bl", "bill of lading", "bol", "bl no", "bl_no", "bl number"],
        "vessel": ["vessel", "vessel name", "ship", "carrier vessel"],
        "voyage": ["voyage", "voyage no", "voy", "voyage number"],
        "description": ["description", "desc", "goods description", "item description", "cargo description", "commodity"],
        "sku": ["sku", "item code", "product code", "part number", "article no", "article number", "item no"],
        "quantity": ["quantity", "qty", "pcs", "units", "total qty", "total quantity"],
        "unit": ["unit", "uom", "unit of measure"],
        "weight_kg": ["weight", "weight_kg", "gross weight", "net weight", "wt", "weight (kg)", "kg"],
        "carton_count": ["carton", "cartons", "ctns", "ctn", "packages", "pkgs", "carton count", "no of cartons"],
    }

    def parse(self, file_path: str, file_type: str = "auto") -> dict[str, Any]:
        path = Path(file_path)
        if file_type == "auto":
            suffix = path.suffix.lower()
            if suffix == ".pdf":
                file_type = "pdf"
            elif suffix in (".xlsx", ".xls"):
                file_type = "excel"
            elif suffix == ".csv":
                file_type = "csv"
            else:
                raise ValueError(f"Unsupported file format: {suffix}")

        parsers = {"pdf": self.parse_pdf, "excel": self.parse_excel, "csv": self.parse_csv}
        if file_type not in parsers:
            raise ValueError(f"Unknown file type: {file_type}")

        return parsers[file_type](file_path)

    def parse_pdf(self, file_path: str) -> dict[str, Any]:
        import pdfplumber

        text_lines: list[str] = []
        tables: list[list[list[str | None]]] = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_lines.extend(page_text.splitlines())
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)

        header = self._extract_header_from_text(text_lines)
        items = self._extract_items_from_tables(tables) if tables else self._extract_items_from_text(text_lines)

        return {**header, "items": items, "source_format": "pdf"}

    def parse_excel(self, file_path: str) -> dict[str, Any]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
        wb.close()

        header = self._extract_header_from_rows(rows)

        header_row_idx = self._find_header_row(rows)
        if header_row_idx is not None:
            col_map = self._map_columns(rows[header_row_idx])
            items = self._parse_item_rows(rows[header_row_idx + 1 :], col_map)
        else:
            items = []

        return {**header, "items": items, "source_format": "excel"}

    def parse_csv(self, file_path: str) -> dict[str, Any]:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            content = f.read()

        reader = csv.reader(io.StringIO(content))
        rows = [[cell.strip() for cell in row] for row in reader]

        header = self._extract_header_from_rows(rows)
        header_row_idx = self._find_header_row(rows)

        if header_row_idx is not None:
            col_map = self._map_columns(rows[header_row_idx])
            items = self._parse_item_rows(rows[header_row_idx + 1 :], col_map)
        else:
            items = []

        return {**header, "items": items, "source_format": "csv"}

    # ── Header extraction ──────────────────────────────────────────────────

    def _extract_header_from_text(self, lines: list[str]) -> dict[str, str | None]:
        header: dict[str, str | None] = {"bl_number": None, "vessel": None, "voyage": None}

        bl_pattern = re.compile(r"(?:b/?l|bill\s+of\s+lading|bol)\s*(?:no\.?|number|#)?\s*[:\-]?\s*([A-Z0-9\-/]+)", re.I)
        vessel_pattern = re.compile(r"vessel\s*(?:name)?\s*[:\-]?\s*(.+)", re.I)
        voyage_pattern = re.compile(r"voyage?\s*(?:no\.?)?\s*[:\-]?\s*([A-Z0-9\-/]+)", re.I)

        for line in lines[:30]:
            if not header["bl_number"]:
                m = bl_pattern.search(line)
                if m:
                    header["bl_number"] = m.group(1).strip()
            if not header["vessel"]:
                m = vessel_pattern.search(line)
                if m:
                    header["vessel"] = m.group(1).strip()
            if not header["voyage"]:
                m = voyage_pattern.search(line)
                if m:
                    header["voyage"] = m.group(1).strip()

        return header

    def _extract_header_from_rows(self, rows: list[list[str]]) -> dict[str, str | None]:
        header: dict[str, str | None] = {"bl_number": None, "vessel": None, "voyage": None}

        for row in rows[:20]:
            joined = " ".join(row).strip()
            if not joined:
                continue
            for i, cell in enumerate(row):
                cell_lower = cell.lower().strip()
                if any(alias in cell_lower for alias in self.FIELD_ALIASES["bl_number"]) and i + 1 < len(row) and row[i + 1]:
                    header["bl_number"] = header["bl_number"] or row[i + 1]
                elif any(alias in cell_lower for alias in ["vessel", "vessel name", "ship"]) and i + 1 < len(row) and row[i + 1]:
                    header["vessel"] = header["vessel"] or row[i + 1]
                elif any(alias in cell_lower for alias in ["voyage", "voy"]) and i + 1 < len(row) and row[i + 1]:
                    header["voyage"] = header["voyage"] or row[i + 1]

        return header

    # ── Item extraction ────────────────────────────────────────────────────

    def _find_header_row(self, rows: list[list[str]]) -> int | None:
        item_keywords = {"description", "desc", "sku", "quantity", "qty", "weight", "carton"}
        for idx, row in enumerate(rows):
            cells_lower = {c.lower().strip() for c in row if c}
            if len(cells_lower & item_keywords) >= 2:
                return idx
        return None

    def _map_columns(self, header_row: list[str]) -> dict[str, int]:
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            cell_lower = cell.lower().strip()
            for field_name, aliases in self.FIELD_ALIASES.items():
                if field_name in ("bl_number", "vessel", "voyage"):
                    continue
                if cell_lower in aliases or any(a in cell_lower for a in aliases):
                    col_map[field_name] = idx
                    break
        return col_map

    def _parse_item_rows(self, rows: list[list[str]], col_map: dict[str, int]) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        for row in rows:
            if not any(cell.strip() for cell in row):
                continue

            item: dict[str, Any] = {}
            for field_name, col_idx in col_map.items():
                if col_idx < len(row):
                    raw = row[col_idx].strip()
                    item[field_name] = self._coerce(field_name, raw)

            if item.get("description") or item.get("sku"):
                items.append(self._normalize_item(item))

        return items

    def _extract_items_from_tables(self, tables: list[list[list[str | None]]]) -> list[dict[str, Any]]:
        all_items: list[dict[str, Any]] = []
        for table in tables:
            if not table or len(table) < 2:
                continue
            header_row = [str(c).strip() if c else "" for c in table[0]]
            col_map = self._map_columns(header_row)
            if not col_map:
                continue
            data_rows = [[str(c).strip() if c else "" for c in row] for row in table[1:]]
            all_items.extend(self._parse_item_rows(data_rows, col_map))
        return all_items

    def _extract_items_from_text(self, lines: list[str]) -> list[dict[str, Any]]:
        """Fallback: attempt line-by-line extraction when no tables found."""
        items: list[dict[str, Any]] = []
        qty_pattern = re.compile(r"(\d[\d,.]*)\s*(pcs|units|ctns|kgs?|sets?)", re.I)
        for line in lines:
            m = qty_pattern.search(line)
            if m:
                items.append({
                    "description": line[:m.start()].strip().rstrip("-:"),
                    "sku": None,
                    "quantity": self._parse_number(m.group(1)),
                    "unit": m.group(2).lower(),
                    "weight_kg": None,
                    "carton_count": None,
                })
        return items

    # ── Normalization helpers ──────────────────────────────────────────────

    def _normalize_item(self, item: dict[str, Any]) -> dict[str, Any]:
        return {
            "description": item.get("description") or "",
            "sku": item.get("sku") or None,
            "quantity": self._to_float(item.get("quantity")),
            "unit": (item.get("unit") or "pcs").lower(),
            "weight_kg": self._to_float(item.get("weight_kg")),
            "carton_count": self._to_int(item.get("carton_count")),
        }

    def _coerce(self, field_name: str, raw: str) -> Any:
        if field_name in ("quantity", "weight_kg"):
            return self._parse_number(raw)
        if field_name == "carton_count":
            return self._parse_int(raw)
        return raw

    @staticmethod
    def _parse_number(raw: Any) -> float | None:
        if raw is None:
            return None
        s = str(raw).replace(",", "").strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_int(raw: Any) -> int | None:
        if raw is None:
            return None
        s = str(raw).replace(",", "").strip()
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_float(val: Any) -> float | None:
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_int(val: Any) -> int | None:
        if val is None:
            return None
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None
