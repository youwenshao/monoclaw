"""Comprehensive FX report generator (PDF + Excel)."""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openclaw_shared.database import get_db

logger = logging.getLogger("openclaw.accounting.fx.report")


def generate_fx_report(
    db_path: str | Path,
    output_dir: str | Path,
    period_start: str | None = None,
    period_end: str | None = None,
    fmt: str = "both",
) -> dict[str, str | None]:
    """Generate a comprehensive FX report in PDF and/or Excel format.

    Args:
        db_path: Path to the fx_tracker database.
        output_dir: Directory to write output files.
        period_start: Optional start date filter (YYYY-MM-DD).
        period_end: Optional end date filter (YYYY-MM-DD).
        fmt: "pdf", "excel", or "both".

    Returns:
        Dict with paths to generated files.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    data = _gather_report_data(db_path, period_start, period_end)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    result: dict[str, str | None] = {"pdf": None, "excel": None}

    if fmt in ("pdf", "both"):
        pdf_path = output_dir / f"fx_report_{timestamp}.pdf"
        _generate_pdf(data, pdf_path)
        result["pdf"] = str(pdf_path)

    if fmt in ("excel", "both"):
        xlsx_path = output_dir / f"fx_report_{timestamp}.xlsx"
        _generate_excel(data, xlsx_path)
        result["excel"] = str(xlsx_path)

    logger.info("FX report generated: %s", result)
    return result


def _gather_report_data(
    db_path: str | Path,
    period_start: str | None,
    period_end: str | None,
) -> dict[str, Any]:
    """Collect all data needed for the FX report."""
    date_clause = ""
    params: list[str] = []
    if period_start:
        date_clause += " AND transaction_date >= ?"
        params.append(period_start)
    if period_end:
        date_clause += " AND transaction_date <= ?"
        params.append(period_end)

    with get_db(db_path) as conn:
        transactions = [dict(r) for r in conn.execute(
            f"SELECT * FROM fx_transactions WHERE 1=1{date_clause} ORDER BY transaction_date",
            params,
        ).fetchall()]

        realized = [dict(r) for r in conn.execute(
            f"""SELECT * FROM fx_transactions
                WHERE is_settled = 1 AND realized_gain_loss IS NOT NULL{date_clause}
                ORDER BY settled_date""",
            params,
        ).fetchall()]

        unrealized = [dict(r) for r in conn.execute(
            "SELECT * FROM revaluations ORDER BY period_end_date DESC"
        ).fetchall()]

        rates = [dict(r) for r in conn.execute(
            "SELECT * FROM exchange_rates ORDER BY date DESC LIMIT 100"
        ).fetchall()]

        exposure = [dict(r) for r in conn.execute(
            """SELECT currency, transaction_type,
                      SUM(foreign_amount) as total_foreign, SUM(hkd_amount) as total_hkd
               FROM fx_transactions WHERE is_settled = 0
               GROUP BY currency, transaction_type"""
        ).fetchall()]

    total_realized = sum(r["realized_gain_loss"] for r in realized)
    total_unrealized = sum(r["unrealized_gain_loss"] for r in unrealized)

    return {
        "transactions": transactions,
        "realized": realized,
        "unrealized": unrealized,
        "rates": rates,
        "exposure": exposure,
        "summary": {
            "total_transactions": len(transactions),
            "total_realized_gl": round(total_realized, 2),
            "total_unrealized_gl": round(total_unrealized, 2),
            "net_fx_impact": round(total_realized + total_unrealized, 2),
            "period_start": period_start,
            "period_end": period_end,
            "generated_at": datetime.now().isoformat(),
        },
    }


def _generate_pdf(data: dict[str, Any], output_path: Path) -> None:
    """Generate the PDF report using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("FXTitle", parent=styles["Title"], fontSize=18, spaceAfter=12)
    heading_style = ParagraphStyle("FXHeading", parent=styles["Heading2"], fontSize=13, spaceAfter=8)
    normal_style = styles["Normal"]

    elements: list = []

    elements.append(Paragraph("Foreign Exchange Report", title_style))
    summary = data["summary"]
    period_text = ""
    if summary["period_start"] and summary["period_end"]:
        period_text = f"Period: {summary['period_start']} to {summary['period_end']}"
    elements.append(Paragraph(period_text or f"As of {date.today().isoformat()}", normal_style))
    elements.append(Spacer(1, 10 * mm))

    # Summary box
    elements.append(Paragraph("Summary", heading_style))
    summary_data = [
        ["Total Transactions", str(summary["total_transactions"])],
        ["Total Realized G/L", f"HKD {summary['total_realized_gl']:,.2f}"],
        ["Total Unrealized G/L", f"HKD {summary['total_unrealized_gl']:,.2f}"],
        ["Net FX Impact", f"HKD {summary['net_fx_impact']:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[120 * mm, 50 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.Color(0.1, 0.1, 0.15)),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("BACKGROUND", (1, 0), (1, -1), colors.Color(0.15, 0.15, 0.2)),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.3, 0.3, 0.4)),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8 * mm))

    # Realized G/L table
    if data["realized"]:
        elements.append(Paragraph("Realized Gains / Losses", heading_style))
        header = ["Date", "Currency", "Foreign Amt", "Orig Rate", "Settle Rate", "G/L (HKD)"]
        rows = [header]
        for r in data["realized"][:50]:
            rows.append([
                r.get("settled_date", ""),
                r["currency"],
                f"{r['foreign_amount']:,.2f}",
                f"{r['exchange_rate']:.4f}",
                f"{r.get('settlement_rate', 0):.4f}",
                f"{r['realized_gain_loss']:,.2f}",
            ])
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Unrealized G/L table
    if data["unrealized"]:
        elements.append(Paragraph("Unrealized Gains / Losses", heading_style))
        header = ["Period End", "Currency", "Outstanding", "Closing Rate", "Unrealized G/L"]
        rows = [header]
        for r in data["unrealized"][:50]:
            rows.append([
                r["period_end_date"],
                r["currency"],
                f"{r['outstanding_foreign_amount']:,.2f}",
                f"{r['closing_rate']:.4f}",
                f"{r['unrealized_gain_loss']:,.2f}",
            ])
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

    doc.build(elements)
    logger.info("PDF report written to %s", output_path)


def _generate_excel(data: dict[str, Any], output_path: Path) -> None:
    """Generate the Excel report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary = data["summary"]
    ws_summary.append(["FX Report Summary"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    for label, value in [
        ("Total Transactions", summary["total_transactions"]),
        ("Total Realized G/L (HKD)", summary["total_realized_gl"]),
        ("Total Unrealized G/L (HKD)", summary["total_unrealized_gl"]),
        ("Net FX Impact (HKD)", summary["net_fx_impact"]),
        ("Period Start", summary["period_start"] or "N/A"),
        ("Period End", summary["period_end"] or "N/A"),
        ("Generated", summary["generated_at"]),
    ]:
        ws_summary.append([label, value])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # Transactions sheet
    ws_tx = wb.create_sheet("Transactions")
    tx_headers = [
        "ID", "Date", "Description", "Currency", "Foreign Amt",
        "Rate", "HKD Amt", "Type", "Nature", "Settled", "G/L",
    ]
    ws_tx.append(tx_headers)
    for col_idx, _ in enumerate(tx_headers, 1):
        cell = ws_tx.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for tx in data["transactions"]:
        ws_tx.append([
            tx["id"],
            tx["transaction_date"],
            tx.get("description", ""),
            tx["currency"],
            tx["foreign_amount"],
            tx["exchange_rate"],
            tx["hkd_amount"],
            tx["transaction_type"],
            tx.get("nature", ""),
            "Yes" if tx["is_settled"] else "No",
            tx.get("realized_gain_loss", ""),
        ])

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws_tx.column_dimensions[col].width = 15

    # Realized G/L sheet
    if data["realized"]:
        ws_gl = wb.create_sheet("Realized G_L")
        gl_headers = ["Settled Date", "Currency", "Foreign Amt", "Orig Rate", "Settle Rate", "G/L (HKD)", "Nature"]
        ws_gl.append(gl_headers)
        for col_idx, _ in enumerate(gl_headers, 1):
            cell = ws_gl.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for r in data["realized"]:
            ws_gl.append([
                r.get("settled_date", ""),
                r["currency"],
                r["foreign_amount"],
                r["exchange_rate"],
                r.get("settlement_rate", ""),
                r["realized_gain_loss"],
                r.get("nature", ""),
            ])

    # Unrealized G/L sheet
    if data["unrealized"]:
        ws_ug = wb.create_sheet("Unrealized G_L")
        ug_headers = ["Period End", "Currency", "Outstanding", "Original HKD", "Closing Rate", "Revalued HKD", "Unrealized G/L"]
        ws_ug.append(ug_headers)
        for col_idx, _ in enumerate(ug_headers, 1):
            cell = ws_ug.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for r in data["unrealized"]:
            ws_ug.append([
                r["period_end_date"],
                r["currency"],
                r["outstanding_foreign_amount"],
                r["original_hkd_amount"],
                r["closing_rate"],
                r["revalued_hkd_amount"],
                r["unrealized_gain_loss"],
            ])

    wb.save(str(output_path))
    logger.info("Excel report written to %s", output_path)
