"""Generate bank reconciliation statements in PDF and Excel formats.

Reconciliation statement format:
  Bank balance per statement
  + Deposits in transit (recorded in books, not yet in bank)
  - Outstanding cheques (issued but not yet cleared)
  ± Other adjustments
  = Adjusted bank balance

  Book balance per ledger
  + Bank credits not in books
  - Bank charges not in books
  ± Other adjustments
  = Adjusted book balance

Both adjusted balances should agree.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any


@dataclass
class ReconciliationItem:
    description: str
    amount: float
    entry_id: int | None = None
    category: str = ""


@dataclass
class ReconciliationStatement:
    as_of_date: date
    bank_name: str
    currency: str

    bank_balance: float
    deposits_in_transit: list[ReconciliationItem] = field(default_factory=list)
    outstanding_cheques: list[ReconciliationItem] = field(default_factory=list)
    bank_adjustments: list[ReconciliationItem] = field(default_factory=list)

    book_balance: float = 0.0
    bank_credits_not_in_books: list[ReconciliationItem] = field(default_factory=list)
    bank_charges_not_in_books: list[ReconciliationItem] = field(default_factory=list)
    book_adjustments: list[ReconciliationItem] = field(default_factory=list)

    @property
    def total_deposits_in_transit(self) -> float:
        return sum(i.amount for i in self.deposits_in_transit)

    @property
    def total_outstanding_cheques(self) -> float:
        return sum(i.amount for i in self.outstanding_cheques)

    @property
    def total_bank_adjustments(self) -> float:
        return sum(i.amount for i in self.bank_adjustments)

    @property
    def adjusted_bank_balance(self) -> float:
        return (
            self.bank_balance
            + self.total_deposits_in_transit
            - self.total_outstanding_cheques
            + self.total_bank_adjustments
        )

    @property
    def total_bank_credits_not_in_books(self) -> float:
        return sum(i.amount for i in self.bank_credits_not_in_books)

    @property
    def total_bank_charges_not_in_books(self) -> float:
        return sum(i.amount for i in self.bank_charges_not_in_books)

    @property
    def total_book_adjustments(self) -> float:
        return sum(i.amount for i in self.book_adjustments)

    @property
    def adjusted_book_balance(self) -> float:
        return (
            self.book_balance
            + self.total_bank_credits_not_in_books
            - self.total_bank_charges_not_in_books
            + self.total_book_adjustments
        )

    @property
    def difference(self) -> float:
        return round(self.adjusted_bank_balance - self.adjusted_book_balance, 2)

    def to_dict(self) -> dict[str, Any]:
        return {
            "as_of_date": self.as_of_date.isoformat(),
            "bank_name": self.bank_name,
            "currency": self.currency,
            "bank_balance": self.bank_balance,
            "deposits_in_transit": [{"desc": i.description, "amount": i.amount} for i in self.deposits_in_transit],
            "outstanding_cheques": [{"desc": i.description, "amount": i.amount} for i in self.outstanding_cheques],
            "adjusted_bank_balance": self.adjusted_bank_balance,
            "book_balance": self.book_balance,
            "bank_credits_not_in_books": [{"desc": i.description, "amount": i.amount} for i in self.bank_credits_not_in_books],
            "bank_charges_not_in_books": [{"desc": i.description, "amount": i.amount} for i in self.bank_charges_not_in_books],
            "adjusted_book_balance": self.adjusted_book_balance,
            "difference": self.difference,
        }


def build_reconciliation_statement(
    bank_balance: float,
    book_balance: float,
    unmatched_bank: list[dict[str, Any]],
    unmatched_ledger: list[dict[str, Any]],
    bank_name: str = "Unknown",
    currency: str = "HKD",
    as_of_date: date | None = None,
) -> ReconciliationStatement:
    """Build a reconciliation statement from unmatched items."""
    stmt = ReconciliationStatement(
        as_of_date=as_of_date or date.today(),
        bank_name=bank_name,
        currency=currency,
        bank_balance=bank_balance,
        book_balance=book_balance,
    )

    for le in unmatched_ledger:
        debit = float(le.get("debit", 0) or 0)
        credit = float(le.get("credit", 0) or 0)
        desc = le.get("description", "")

        if credit > 0:
            stmt.deposits_in_transit.append(ReconciliationItem(
                description=desc, amount=credit, entry_id=le.get("id"),
                category="deposit_in_transit",
            ))
        elif debit > 0:
            acct = (le.get("account_code") or "").lower()
            txn_type_hint = le.get("transaction_type", "")
            if "cheque" in desc.lower() or "chq" in desc.lower() or txn_type_hint == "cheque":
                stmt.outstanding_cheques.append(ReconciliationItem(
                    description=desc, amount=debit, entry_id=le.get("id"),
                    category="outstanding_cheque",
                ))
            else:
                stmt.bank_adjustments.append(ReconciliationItem(
                    description=f"Book entry not in bank: {desc}",
                    amount=-debit, entry_id=le.get("id"),
                    category="book_not_in_bank",
                ))

    for b in unmatched_bank:
        debit = float(b.get("debit", 0) or 0)
        credit = float(b.get("credit", 0) or 0)
        desc = b.get("description", "")
        txn_type = (b.get("transaction_type") or "").lower()

        if credit > 0:
            stmt.bank_credits_not_in_books.append(ReconciliationItem(
                description=desc, amount=credit, entry_id=b.get("id"),
                category="bank_credit_not_in_books",
            ))
        elif debit > 0:
            if txn_type in ("charge", "fee") or "fee" in desc.lower() or "charge" in desc.lower():
                stmt.bank_charges_not_in_books.append(ReconciliationItem(
                    description=desc, amount=debit, entry_id=b.get("id"),
                    category="bank_charge",
                ))
            else:
                stmt.book_adjustments.append(ReconciliationItem(
                    description=f"Bank debit not in books: {desc}",
                    amount=-debit, entry_id=b.get("id"),
                    category="bank_not_in_books",
                ))

    return stmt


def generate_pdf_report(
    statement: ReconciliationStatement,
    output_dir: Path,
    firm_name: str = "",
) -> Path:
    """Generate a reconciliation statement PDF using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    filename = f"reconciliation_{statement.bank_name.replace(' ', '_')}_{statement.as_of_date.isoformat()}.pdf"
    output_path = output_dir / filename

    doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    elements.append(Paragraph(firm_name or "Bank Reconciliation Statement", styles["Title"]))
    elements.append(Paragraph(
        f"Bank: {statement.bank_name} | Currency: {statement.currency} | "
        f"As of: {statement.as_of_date.strftime('%d %B %Y')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 10 * mm))

    def _fmt(amount: float) -> str:
        return f"{amount:,.2f}"

    # Bank side
    bank_data = [
        ["Bank Balance per Statement", "", _fmt(statement.bank_balance)],
    ]
    if statement.deposits_in_transit:
        bank_data.append(["Add: Deposits in Transit", "", ""])
        for item in statement.deposits_in_transit:
            bank_data.append(["", item.description, _fmt(item.amount)])
        bank_data.append(["", "Total Deposits in Transit", _fmt(statement.total_deposits_in_transit)])

    if statement.outstanding_cheques:
        bank_data.append(["Less: Outstanding Cheques", "", ""])
        for item in statement.outstanding_cheques:
            bank_data.append(["", item.description, f"({_fmt(item.amount)})"])
        bank_data.append(["", "Total Outstanding Cheques", f"({_fmt(statement.total_outstanding_cheques)})"])

    bank_data.append(["Adjusted Bank Balance", "", _fmt(statement.adjusted_bank_balance)])

    bank_table = Table(bank_data, colWidths=[55 * mm, 70 * mm, 40 * mm])
    bank_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 2, colors.black),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(bank_table)
    elements.append(Spacer(1, 8 * mm))

    # Book side
    book_data = [
        ["Book Balance per Ledger", "", _fmt(statement.book_balance)],
    ]
    if statement.bank_credits_not_in_books:
        book_data.append(["Add: Bank Credits Not in Books", "", ""])
        for item in statement.bank_credits_not_in_books:
            book_data.append(["", item.description, _fmt(item.amount)])
        book_data.append(["", "Total", _fmt(statement.total_bank_credits_not_in_books)])

    if statement.bank_charges_not_in_books:
        book_data.append(["Less: Bank Charges Not in Books", "", ""])
        for item in statement.bank_charges_not_in_books:
            book_data.append(["", item.description, f"({_fmt(item.amount)})"])
        book_data.append(["", "Total", f"({_fmt(statement.total_bank_charges_not_in_books)})"])

    book_data.append(["Adjusted Book Balance", "", _fmt(statement.adjusted_book_balance)])

    book_table = Table(book_data, colWidths=[55 * mm, 70 * mm, 40 * mm])
    book_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 2, colors.black),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(book_table)
    elements.append(Spacer(1, 6 * mm))

    diff = statement.difference
    diff_color = "green" if abs(diff) < 0.01 else "red"
    elements.append(Paragraph(
        f'<b>Difference: <font color="{diff_color}">{_fmt(diff)}</font></b>',
        styles["Normal"],
    ))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    ))

    doc.build(elements)
    return output_path


def generate_excel_report(
    statement: ReconciliationStatement,
    output_dir: Path,
    firm_name: str = "",
) -> Path:
    """Generate a reconciliation statement Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    filename = f"reconciliation_{statement.bank_name.replace(' ', '_')}_{statement.as_of_date.isoformat()}.xlsx"
    output_path = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    currency_fmt = '#,##0.00'
    thin_border = Border(bottom=Side(style="thin"))

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18

    row = 1
    ws.cell(row=row, column=1, value=firm_name or "Bank Reconciliation Statement").font = bold_large
    row += 1
    ws.cell(row=row, column=1, value=f"Bank: {statement.bank_name}")
    ws.cell(row=row, column=2, value=f"Currency: {statement.currency}")
    ws.cell(row=row, column=3, value=f"As of: {statement.as_of_date.isoformat()}")
    row += 2

    # Bank side
    ws.cell(row=row, column=1, value="Bank Balance per Statement").font = bold
    ws.cell(row=row, column=3, value=statement.bank_balance).number_format = currency_fmt
    row += 1

    if statement.deposits_in_transit:
        ws.cell(row=row, column=1, value="Add: Deposits in Transit").font = bold
        row += 1
        for item in statement.deposits_in_transit:
            ws.cell(row=row, column=2, value=item.description)
            ws.cell(row=row, column=3, value=item.amount).number_format = currency_fmt
            row += 1
        ws.cell(row=row, column=2, value="Subtotal").font = bold
        ws.cell(row=row, column=3, value=statement.total_deposits_in_transit).number_format = currency_fmt
        row += 1

    if statement.outstanding_cheques:
        ws.cell(row=row, column=1, value="Less: Outstanding Cheques").font = bold
        row += 1
        for item in statement.outstanding_cheques:
            ws.cell(row=row, column=2, value=item.description)
            ws.cell(row=row, column=3, value=-item.amount).number_format = currency_fmt
            row += 1
        ws.cell(row=row, column=2, value="Subtotal").font = bold
        ws.cell(row=row, column=3, value=-statement.total_outstanding_cheques).number_format = currency_fmt
        row += 1

    ws.cell(row=row, column=1, value="Adjusted Bank Balance").font = bold
    c = ws.cell(row=row, column=3, value=statement.adjusted_bank_balance)
    c.number_format = currency_fmt
    c.font = bold
    c.border = thin_border
    row += 2

    # Book side
    ws.cell(row=row, column=1, value="Book Balance per Ledger").font = bold
    ws.cell(row=row, column=3, value=statement.book_balance).number_format = currency_fmt
    row += 1

    if statement.bank_credits_not_in_books:
        ws.cell(row=row, column=1, value="Add: Bank Credits Not in Books").font = bold
        row += 1
        for item in statement.bank_credits_not_in_books:
            ws.cell(row=row, column=2, value=item.description)
            ws.cell(row=row, column=3, value=item.amount).number_format = currency_fmt
            row += 1

    if statement.bank_charges_not_in_books:
        ws.cell(row=row, column=1, value="Less: Bank Charges Not in Books").font = bold
        row += 1
        for item in statement.bank_charges_not_in_books:
            ws.cell(row=row, column=2, value=item.description)
            ws.cell(row=row, column=3, value=-item.amount).number_format = currency_fmt
            row += 1

    ws.cell(row=row, column=1, value="Adjusted Book Balance").font = bold
    c = ws.cell(row=row, column=3, value=statement.adjusted_book_balance)
    c.number_format = currency_fmt
    c.font = bold
    c.border = thin_border
    row += 2

    ws.cell(row=row, column=1, value="Difference").font = bold
    diff_cell = ws.cell(row=row, column=3, value=statement.difference)
    diff_cell.number_format = currency_fmt
    diff_cell.font = Font(bold=True, color="FF0000" if abs(statement.difference) >= 0.01 else "008000")

    wb.save(str(output_path))
    return output_path
