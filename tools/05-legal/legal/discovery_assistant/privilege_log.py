"""Generate HK High Court compliant privilege log as Excel."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from openclaw_shared.database import get_db

LOG_COLUMNS = [
    "No.",
    "Date",
    "Document Type",
    "Author",
    "Recipients",
    "Description",
    "Privilege Claimed",
]

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

PRIVILEGE_LABEL_MAP = {
    "legal_professional_privilege": "Legal Professional Privilege (solicitor-client)",
    "litigation_privilege": "Litigation Privilege",
    "without_prejudice": "Without Prejudice Privilege",
    "general_privilege": "Legal Professional Privilege",
}


def _fetch_privileged_documents(db_path: str | Path) -> list[dict[str, Any]]:
    """Retrieve all documents classified as privileged or partial."""
    with get_db(db_path) as conn:
        rows = conn.execute(
            """
            SELECT
                d.id,
                d.date_created,
                d.doc_type,
                d.author,
                d.recipients,
                d.subject,
                c.privilege_status,
                c.privilege_type,
                COALESCE(pl.description, d.subject) AS description,
                COALESCE(pl.privilege_basis, c.privilege_type) AS privilege_basis
            FROM documents d
            JOIN classifications c ON c.document_id = d.id
            LEFT JOIN privilege_log pl ON pl.document_id = d.id
            WHERE c.privilege_status IN ('privileged', 'partial')
            ORDER BY d.date_created ASC
            """,
        ).fetchall()
    return [dict(r) for r in rows]


def generate_privilege_log(db_path: str | Path, output_path: str | Path) -> Path:
    """Generate a HK High Court compliant privilege log as an Excel workbook.

    Returns the Path to the written file.
    """
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    docs = _fetch_privileged_documents(db_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Privilege Log"

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "PRIVILEGE LOG"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    subtitle = ws["A2"]
    subtitle.value = f"Generated: {date.today().strftime('%d %B %Y')}"
    subtitle.font = Font(name="Arial", size=10, italic=True, color="666666")
    subtitle.alignment = Alignment(horizontal="center")

    header_row = 4
    for col_idx, col_name in enumerate(LOG_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    col_widths = [6, 14, 16, 25, 30, 45, 35]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    for row_idx, doc in enumerate(docs, start=1):
        excel_row = header_row + row_idx

        doc_type_label = {
            "email": "Email",
            "attachment": "Attachment",
            "standalone": "Document",
        }.get(doc.get("doc_type", ""), doc.get("doc_type", ""))

        privilege_label = PRIVILEGE_LABEL_MAP.get(
            doc.get("privilege_basis", ""),
            doc.get("privilege_basis") or "Legal Professional Privilege",
        )
        if doc.get("privilege_status") == "partial":
            privilege_label = f"Partial — {privilege_label}"

        values = [
            row_idx,
            doc.get("date_created", ""),
            doc_type_label,
            doc.get("author", ""),
            doc.get("recipients", ""),
            doc.get("description", ""),
            privilege_label,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr  # noqa: B017
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = f"{header_row}:{header_row}"

    wb.save(str(output))
    return output
